import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_option('deprecation.showPyplotGlobalUse', False)

# Set page title and background color
st.set_page_config(page_title="Zimmerpflanzenwächter", page_icon="🌱", layout="wide")
st.title('Zimmerpflanzenwächter')

# Funktionen zum Laden von Daten und Anzeigen von Inhalten definieren
@st.cache_resource
#def load_data():
    #return pd.read_excel("/hdd/Datenbank/Pflanzen_Daten.xlsx")

def display_content(tab_index):
    st.write(f"Inhalt für Reiter {tab_index}")

# Seitenleiste mit Buttons für jeden Reiter erstellen
st.sidebar.write('# Pflanzen')  # Header für die Seitenleiste

# Seitenleiste mit Buttons für jeden Reiter erstellen
selected_tab_1 = st.sidebar.button("Pflanze 1")
selected_tab_2 = st.sidebar.button("Pflanze 2")
selected_tab_3 = st.sidebar.button("Pflanze 3")
selected_tab_4 = st.sidebar.button("Pflanze 4")

selected_tab = None
if selected_tab_1:
    selected_tab = "Pflanze 1"
elif selected_tab_2:
    selected_tab = "Pflanze 2"
elif selected_tab_3:
    selected_tab = "Pflanze 3"
elif selected_tab_4:
    selected_tab = "Pflanze 4"

# Alle anderen Reiter anzeigen, wenn ein Reiter ausgewählt ist
if selected_tab:
    selected_tab_1 = True
    selected_tab_2 = True
    selected_tab_3 = True
    selected_tab_4 = True

# Hauptinhalt basierend auf ausgewähltem Reiter anzeigen
if selected_tab == "Pflanze 1":
    
    st.subheader("Monstera")
    st.write("Die Pflanze steht im Wohnzimmer")
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')
    sheet = book['Pflanze1']

    # Daten auslesen und in DataFrame konvertieren
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
    df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])

    # Konvertieren der Bodenfeuchtigkeitswerte in numerische Datentypen
    df['Bodenfeuchtigkeit'] = pd.to_numeric(df['Bodenfeuchtigkeit'], errors='coerce')

    # Formatieren der Uhrzeit-Spalte, um die Sekunden auszublenden
    df['Uhrzeit'] = pd.to_datetime(df['Uhrzeit']).dt.strftime('%H:%M')

    # Linien Diagramm erstellen mit Seaborn
    plt.figure(figsize=(10, 5))
    sns.lineplot(data=df, x='Uhrzeit', y='Bodenfeuchtigkeit', marker='o', linestyle='-')
    plt.xlabel('Uhrzeit')
    plt.ylabel('Bodenfeuchtigkeit (%)')
    plt.title('Bodenfeuchtigkeitswerte für Monstera')
    plt.xticks(rotation=45)

    # Skalierung der Y-Achse festlegen
    plt.ylim(0, 100)

    # Diagramm im Streamlit anzeigen
    st.pyplot()
    st.button("Zurück zur Standardansicht")

elif selected_tab == "Pflanze 2":

    st.subheader("Pflanze 2")
    st.write("Die Pflanze steht im X")

    # Laden der Daten aus der Excel-Datei
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')
    sheet = book['Pflanze2']

    # Daten auslesen und in DataFrame konvertieren
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
    df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])

    # Konvertieren der Bodenfeuchtigkeitswerte in numerische Datentypen
    df['Bodenfeuchtigkeit'] = pd.to_numeric(df['Bodenfeuchtigkeit'], errors='coerce')

    # Formatieren der Uhrzeit-Spalte, um die Sekunden auszublenden
    df['Uhrzeit'] = pd.to_datetime(df['Uhrzeit']).dt.strftime('%H:%M')

    # Linien Diagramm erstellen mit Seaborn
    plt.figure(figsize=(10, 5))
    sns.lineplot(data=df, x='Uhrzeit', y='Bodenfeuchtigkeit', marker='o', linestyle='-')
    plt.xlabel('Uhrzeit')
    plt.ylabel('Bodenfeuchtigkeit (%)')
    plt.title('Bodenfeuchtigkeitswerte für Pflanze 2')
    plt.xticks(rotation=45)

    # Skalierung der Y-Achse festlegen
    plt.ylim(0, 100)

    # Diagramm im Streamlit anzeigen
    st.pyplot()

    st.button("Zurück zur Standardansicht")


elif selected_tab == "Pflanze 3":

    st.subheader("Pflanze 3")
    st.write("Die Pflanze steht im X")
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')
    sheet = book['Pflanze3']

    # Daten auslesen und in DataFrame konvertieren
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
    df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])

    # Konvertieren der Bodenfeuchtigkeitswerte in numerische Datentypen
    df['Bodenfeuchtigkeit'] = pd.to_numeric(df['Bodenfeuchtigkeit'], errors='coerce')

    # Formatieren der Uhrzeit-Spalte, um die Sekunden auszublenden
    df['Uhrzeit'] = pd.to_datetime(df['Uhrzeit']).dt.strftime('%H:%M')

    # Linien Diagramm erstellen mit Seaborn
    plt.figure(figsize=(10, 5))
    sns.lineplot(data=df, x='Uhrzeit', y='Bodenfeuchtigkeit', marker='o', linestyle='-')
    plt.xlabel('Uhrzeit')
    plt.ylabel('Bodenfeuchtigkeit (%)')
    plt.title('Bodenfeuchtigkeitswerte für Pflanze 3')
    plt.xticks(rotation=45)

    # Skalierung der Y-Achse festlegen
    plt.ylim(0, 100)

    # Diagramm im Streamlit anzeigen
    st.pyplot()
    st.button("Zurück zur Standardansicht")


elif selected_tab == "Pflanze 4":

    st.subheader("Pflanze 4")
    st.write("Die Pflanze steht im X")
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')
    sheet = book['Pflanze4']
    data = sheet.values
   # Daten auslesen und in DataFrame konvertieren
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
    df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])

    # Konvertieren der Bodenfeuchtigkeitswerte in numerische Datentypen
    df['Bodenfeuchtigkeit'] = pd.to_numeric(df['Bodenfeuchtigkeit'], errors='coerce')

    # Formatieren der Uhrzeit-Spalte, um die Sekunden auszublenden
    df['Uhrzeit'] = pd.to_datetime(df['Uhrzeit']).dt.strftime('%H:%M')

    # Linien Diagramm erstellen mit Seaborn
    plt.figure(figsize=(10, 5))
    sns.lineplot(data=df, x='Uhrzeit', y='Bodenfeuchtigkeit', marker='o', linestyle='-')
    plt.xlabel('Uhrzeit')
    plt.ylabel('Bodenfeuchtigkeit (%)')
    plt.title('Bodenfeuchtigkeitswerte für Pflanze 4')
    plt.xticks(rotation=45)

    # Skalierung der Y-Achse festlegen
    plt.ylim(0, 100)

    # Diagramm im Streamlit anzeigen
    st.pyplot()
    st.button("Zurück zur Standardansicht")


else:
    #st.image('/hdd/Datenbank/Bild_Webserver.jpg')
    st.markdown("<h3>Willkommen zum Zimmerpflanzenwächter!</h3>", unsafe_allow_html=True)
    st.write( """ 
             
             Hier werden die aktuellen Bodenfeuchtigkeitswerte deiner Pflanzen visualisiert. 
             Der Gartenwächter überwacht kontinuierlich die Feuchtigkeit im Boden und stellt sicher, dass deine Pflanzen optimal bewässert werden.
             """)
    
        # Laden der Daten aus der Excel-Datei
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')

    # Streamlit-App
    st.title('Bodenfeuchtigkeitswerte für Pflanzen')

    # Platzhalter für das Plot
    plt.figure(figsize=(10, 5))

    # Daten für jede Pflanze plotten
    for plant_name in ['Pflanze1', 'Pflanze2', 'Pflanze3', 'Pflanze4']:
        sheet = book[plant_name]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
        df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])

        # Konvertieren der Bodenfeuchtigkeitswerte in numerische Datentypen
        df['Bodenfeuchtigkeit'] = pd.to_numeric(df['Bodenfeuchtigkeit'], errors='coerce')

        # Linien Diagramm erstellen mit Seaborn
        sns.lineplot(data=df, x='Uhrzeit', y='Bodenfeuchtigkeit', marker='o', linestyle='-', label=plant_name)

    plt.xlabel('Uhrzeit')
    plt.ylabel('Bodenfeuchtigkeit (%)')
    plt.title('Bodenfeuchtigkeitswerte für alle Pflanzen')
    plt.xticks([])
    plt.ylim(0, 100)
    plt.legend()

    # Diagramm im Streamlit anzeigen
    st.pyplot()

    

 # Laden der Daten aus der Excel-Datei für alle Pflanzen
    book = load_workbook('/hdd/Datenbank/Pflanzen_Daten.xlsx')

    # Eine leere Liste für die Daten aller Pflanzen erstellen
    all_data = []

    # Liste aller Pflanzen in der gewünschten Reihenfolge
    all_plants = ["Pflanze1", "Pflanze2", "Pflanze3", "Pflanze4"]

    # Durch alle Blätter in der Excel-Datei iterieren und Daten für jede Pflanze hinzufügen
    for sheet_name in all_plants:
        sheet = book[sheet_name]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)]
        df = pd.DataFrame(data, columns=['Uhrzeit', 'Bodenfeuchtigkeit'])
        df['Pflanze'] = sheet_name  # Eine Spalte für die Pflanzen hinzufügen
        all_data.append(df)

    # Daten aller Pflanzen zusammenführen
    all_data_df = pd.concat(all_data)

    # Tabelle im Streamlit anzeigen
    st.subheader("Alle Pflanzen - Bodenfeuchtigkeitswerte")

    # Umstrukturierung der Daten für die Anzeige und Ersatz von NaN-Werten
    pivot_table = all_data_df.pivot_table(index='Uhrzeit', columns='Pflanze', values='Bodenfeuchtigkeit', aggfunc='first').fillna('Nicht verfügbar')

    # Spalten in der gewünschten Reihenfolge anordnen
    pivot_table = pivot_table.reindex(columns=all_plants)

    # Anzeige der umstrukturierten Tabelle
    st.write(pivot_table, width=1200)
